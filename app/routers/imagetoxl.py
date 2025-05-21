from fastapi import APIRouter, File, UploadFile, Form, HTTPException, BackgroundTasks, Query, Path
from fastapi.responses import FileResponse
from pathlib import Path as FilePath
import cv2
import numpy as np
import pandas as pd
import pytesseract
from PIL import Image
import os
import uuid
import shutil
import json
from typing import List, Dict, Optional, Any
import tempfile
from pydantic import BaseModel
from openpyxl import Workbook
import time
import logging
import csv

# Additional OCR and image processing libraries
import easyocr
import tabula
import camelot
import pdfplumber
from paddleocr import PaddleOCR

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("image-to-excel-system")

# Suppress PaddleOCR logs below ERROR
logging.getLogger("ppocr").setLevel(logging.ERROR)
logging.getLogger("paddleocr").setLevel(logging.ERROR)

# Create the FastAPI app
router = APIRouter(
    prefix="/imagetoxl",
    tags=["imagetoxl"],
    responses={404: {"description": "Not found"}},
)

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Create directories for storing files
BASE_DIR = FilePath("./storage")
UPLOADS_DIR = BASE_DIR / "uploads"
RESULTS_DIR = BASE_DIR / "results"
TEMP_DIR = BASE_DIR / "temp"

for directory in [BASE_DIR, UPLOADS_DIR, RESULTS_DIR, TEMP_DIR]:
    directory.mkdir(exist_ok=True, parents=True)
    # Create user specific directories as needed

# Pydantic models
class ExtractionResponse(BaseModel):
    task_id: str
    message: str
    status: str

class ExtractionResult(BaseModel):
    task_id: str
    user_id: str
    original_filename: str
    extraction_method: str
    excel_path: Optional[str] = None
    csv_path: Optional[str] = None
    preview_data: List[Dict[str, Any]]
    status: str
    error: Optional[str] = None
    processing_time: float

# Initialize OCR engines
try:
    # Initialize EasyOCR reader
    easyocr_reader = easyocr.Reader(['en'])
    
    # Initialize PaddleOCR
    paddle_ocr = PaddleOCR(use_angle_cls=True, lang='en')
    
    # Set Tesseract path if needed
    # pytesseract.pytesseract.tesseract_cmd = r'/usr/bin/tesseract'
except Exception as e:
    logger.error(f"Error initializing OCR engines: {str(e)}")
    # Continue, individual OCR methods will handle failures

# Helper functions
def get_user_dir(user_id: str):
    """Create and return user-specific directories"""
    user_uploads = UPLOADS_DIR / user_id
    user_results = RESULTS_DIR / user_id
    user_temp = TEMP_DIR / user_id
    
    for directory in [user_uploads, user_results, user_temp]:
        directory.mkdir(exist_ok=True, parents=True)
    
    return user_uploads, user_results, user_temp

def save_to_csv(data: List[List[Any]], filepath: str):
    """Save data to CSV file"""
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        for row in data:
            writer.writerow(row)
    return filepath

def preprocess_image(image_path: str, enhance_resolution: bool = False, denoise: bool = False,
                    threshold: bool = False, deskew: bool = False):
    """Preprocess image to improve OCR quality"""
    # Read the image
    img = cv2.imread(image_path)
    if img is None:
        raise ValueError("Could not read image")
    
    # Apply preprocessing steps based on parameters
    if enhance_resolution:
        # Upscale low-resolution images
        img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    
    if denoise:
        # Remove noise
        img = cv2.fastNlMeansDenoisingColored(img, None, 10, 10, 7, 21)
    
    if threshold:
        # Convert to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # Apply adaptive thresholding
        img = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                   cv2.THRESH_BINARY, 11, 2)
    
    if deskew:
        # Deskew the image if it's skewed
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img
        coords = np.column_stack(np.where(gray > 0))
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
        (h, w) = img.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        img = cv2.warpAffine(img, M, (w, h), flags=cv2.INTER_CUBIC, 
                            borderMode=cv2.BORDER_REPLICATE)
    
    # Save the preprocessed image
    preprocessed_path = image_path.replace('.', '_preprocessed.')
    cv2.imwrite(preprocessed_path, img)
    return preprocessed_path

# 1. Basic Pytesseract OCR approach
def extract_with_tesseract(image_path: str, preprocess: bool = True):
    """Extract data using Tesseract OCR"""
    try:
        if preprocess:
            image_path = preprocess_image(image_path, enhance_resolution=True, denoise=True, threshold=True)
        
        img = Image.open(image_path)
        # Use pytesseract with specific configuration for table detection
        custom_config = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
        text = pytesseract.image_to_string(img, config=custom_config)
        
        # Get bounding boxes of detected text (for table structure reconstruction)
        boxes = pytesseract.image_to_data(img, config=custom_config, output_type=pytesseract.Output.DICT)
        
        # Process the text line by line to create a structured table
        lines = text.split('\n')
        lines = [line.strip() for line in lines if line.strip()]
        
        # Try to detect columns by analyzing consistent spaces
        data = []
        for line in lines:
            # Split by multiple spaces
            cells = [cell.strip() for cell in line.split('  ') if cell.strip()]
            data.append(cells)
        
        return data
    except Exception as e:
        logger.error(f"Tesseract extraction error: {str(e)}")
        return [["Error in Tesseract extraction", str(e)]]

# 2. EasyOCR approach
def extract_with_easyocr(image_path: str, preprocess: bool = True):
    """Extract data using EasyOCR"""
    try:
        if preprocess:
            image_path = preprocess_image(image_path, enhance_resolution=True, denoise=True)
        
        # Perform text detection
        results = easyocr_reader.readtext(image_path)
        
        # Group text by y-coordinate (approximate rows)
        # Sort by y-coordinate first
        results.sort(key=lambda x: (int(sum([pt[1] for pt in x[0]]) / 4), int(sum([pt[0] for pt in x[0]]) / 4)))
        
        row_threshold = 10  # pixels, for grouping text in the same row
        current_row = []
        rows = []
        last_y = -100
        
        for bbox, text, conf in results:
            # Calculate center y-coordinate
            center_y = sum([pt[1] for pt in bbox]) / 4
            
            # If this text is on a new row
            if abs(center_y - last_y) > row_threshold and current_row:
                # Sort current row by x-coordinate
                current_row.sort(key=lambda x: x[0])
                # Add only the text to the rows
                rows.append([cell[1] for cell in current_row])
                current_row = []
            
            # Calculate center x-coordinate
            center_x = sum([pt[0] for pt in bbox]) / 4
            current_row.append((center_x, text, conf))
            last_y = center_y
        
        # Don't forget the last row
        if current_row:
            current_row.sort(key=lambda x: x[0])
            rows.append([cell[1] for cell in current_row])
        
        return rows
    except Exception as e:
        logger.error(f"EasyOCR extraction error: {str(e)}")
        return [["Error in EasyOCR extraction", str(e)]]

# 3. PaddleOCR approach
def extract_with_paddleocr(image_path: str, preprocess: bool = True):
    """Extract data using PaddleOCR"""
    try:
        if preprocess:
            image_path = preprocess_image(image_path, enhance_resolution=True, denoise=True)
        
        # Perform OCR
        result = paddle_ocr.ocr(image_path, cls=True)
        
        # Group by rows (similar y-coordinates)
        if result:
            # Sort by y-coordinate
            flat_results = []
            for line in result:
                for item in line:
                    bbox, (text, conf) = item
                    # Calculate center coordinates
                    center_x = (bbox[0][0] + bbox[1][0] + bbox[2][0] + bbox[3][0]) / 4
                    center_y = (bbox[0][1] + bbox[1][1] + bbox[2][1] + bbox[3][1]) / 4
                    flat_results.append((center_x, center_y, text, conf))
            
            # Sort by y-coordinate
            flat_results.sort(key=lambda x: x[1])
            
            row_threshold = 10  # pixels
            current_row = []
            rows = []
            last_y = -100
            
            for center_x, center_y, text, conf in flat_results:
                # If this text is on a new row
                if abs(center_y - last_y) > row_threshold and current_row:
                    # Sort current row by x-coordinate
                    current_row.sort(key=lambda x: x[0])
                    # Add only the text to the rows
                    rows.append([cell[2] for cell in current_row])
                    current_row = []
                
                current_row.append((center_x, center_y, text, conf))
                last_y = center_y
            
            # Don't forget the last row
            if current_row:
                current_row.sort(key=lambda x: x[0])
                rows.append([cell[2] for cell in current_row])
            
            return rows
        return [["No text detected"]]
    except Exception as e:
        logger.error(f"PaddleOCR extraction error: {str(e)}")
        return [["Error in PaddleOCR extraction", str(e)]]

# 4. PDF Table extraction approach (for PDF files or images converted to PDF)
def extract_from_pdf_tables(image_path: str, convert_to_pdf: bool = True):
    """Extract tables from PDF or convert image to PDF first"""
    try:
        pdf_path = image_path
        
        # If the input is an image, convert it to PDF first
        if convert_to_pdf and not image_path.lower().endswith('.pdf'):
            pdf_path = image_path.replace('.', '_converted.') + '.pdf'
            img = Image.open(image_path)
            img.save(pdf_path, 'PDF')
        
        # Try different PDF table extraction methods
        
        # 1. Try with tabula-py
        try:
            tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
            if tables and len(tables) > 0:
                # Convert the first table to list of lists
                df = tables[0]
                return [df.columns.tolist()] + df.values.tolist()
        except Exception as e:
            logger.warning(f"Tabula extraction failed: {str(e)}")
        
        # 2. Try with camelot
        try:
            tables = camelot.read_pdf(pdf_path)
            if tables and len(tables) > 0:
                df = tables[0].df
                return df.values.tolist()
        except Exception as e:
            logger.warning(f"Camelot extraction failed: {str(e)}")
        
        # 3. Try with pdfplumber
        try:
            with pdfplumber.open(pdf_path) as pdf:
                first_page = pdf.pages[0]
                tables = first_page.extract_tables()
                if tables and len(tables) > 0:
                    return tables[0]
        except Exception as e:
            logger.warning(f"PDFPlumber extraction failed: {str(e)}")
        
        # If all methods failed
        return [["No tables detected in PDF"]]
    except Exception as e:
        logger.error(f"PDF table extraction error: {str(e)}")
        return [["Error in PDF table extraction", str(e)]]

# 5. Computer vision based approach for structured tables
def extract_table_with_cv(image_path: str):
    """Extract table structure using computer vision techniques"""
    try:
        # Read the image
        img = cv2.imread(image_path)
        
        # Preprocess the image
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                      cv2.THRESH_BINARY_INV, 11, 2)
        
        # Detect horizontal lines
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
        horizontal_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=1)
        
        # Detect vertical lines
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
        vertical_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=1)
        
        # Combine lines
        table_structure = cv2.add(horizontal_lines, vertical_lines)
        
        # Find contours
        contours, _ = cv2.findContours(table_structure, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        
        # Find table grid cells
        bounding_boxes = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            if w > 20 and h > 20 and w < img.shape[1] * 0.9 and h < img.shape[0] * 0.9:
                bounding_boxes.append((x, y, w, h))
        
        # Sort cells by position (rows then columns)
        bounding_boxes.sort(key=lambda b: (b[1], b[0]))
        
        # Group cells into rows
        row_threshold = 10
        current_row = []
        rows = []
        last_y = -100
        
        for x, y, w, h in bounding_boxes:
            if abs(y - last_y) > row_threshold and current_row:
                # Sort current row by x-coordinate
                current_row.sort(key=lambda cell: cell[0])
                rows.append(current_row)
                current_row = []
            
            current_row.append((x, y, w, h))
            last_y = y
        
        # Add the last row
        if current_row:
            current_row.sort(key=lambda cell: cell[0]) 
            rows.append(current_row)
        
        # Extract text from each cell
        table_data = []
        for row in rows:
            row_data = []
            for x, y, w, h in row:
                # Extract the cell image
                cell_image = gray[y:y+h, x:x+w]
                
                # Perform OCR on the cell
                text = pytesseract.image_to_string(cell_image, config='--psm 7').strip()
                row_data.append(text)
            
            table_data.append(row_data)
        
        return table_data
    except Exception as e:
        logger.error(f"CV-based table extraction error: {str(e)}")
        return [["Error in CV-based table extraction", str(e)]]

# 6. Ensemble approach - combines multiple methods and picks the best result
def extract_with_ensemble(image_path: str):
    """Apply multiple extraction methods and select the best result"""
    results = {}
    
    # Apply all extraction methods
    try:
        results["tesseract"] = extract_with_tesseract(image_path)
    except Exception as e:
        logger.error(f"Tesseract failed in ensemble: {str(e)}")
    
    try:
        results["easyocr"] = extract_with_easyocr(image_path)
    except Exception as e:
        logger.error(f"EasyOCR failed in ensemble: {str(e)}")
    
    try:
        results["paddleocr"] = extract_with_paddleocr(image_path)
    except Exception as e:
        logger.error(f"PaddleOCR failed in ensemble: {str(e)}")
    
    try:
        # Only try CV approach if the image isn't a PDF
        if not image_path.lower().endswith('.pdf'):
            results["cv_table"] = extract_table_with_cv(image_path)
    except Exception as e:
        logger.error(f"CV table extraction failed in ensemble: {str(e)}")
    
    try:
        # Convert to PDF and extract if not already a PDF
        if not image_path.lower().endswith('.pdf'):
            results["pdf_table"] = extract_from_pdf_tables(image_path)
    except Exception as e:
        logger.error(f"PDF table extraction failed in ensemble: {str(e)}")
    
    # Select the best result by a simple heuristic: most rows and columns
    best_method = None
    max_cells = 0
    
    for method, data in results.items():
        # Calculate total number of cells
        num_cells = sum(len(row) for row in data) if data else 0
        num_rows = len(data) if data else 0
        
        # If this method found more cells, it's potentially better
        if num_cells > max_cells and num_rows > 1:  # Ensure we have at least 2 rows
            max_cells = num_cells
            best_method = method
    
    if best_method:
        logger.info(f"Selected best method: {best_method} with {max_cells} cells")
        return results[best_method], best_method
    else:
        # If all methods failed or found nothing substantial, return empty result
        return [["No data extracted"]], "none"

# Background task for processing images
async def process_image(user_id: str, file_path: str, original_filename: str, task_id: str, extraction_method: str):
    start_time = time.time()
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    result = {
        "task_id": task_id,
        "user_id": user_id,
        "original_filename": original_filename,
        "extraction_method": extraction_method,
        "preview_data": [],
        "status": "processing",
        "error": None,
        "processing_time": 0
    }
    
    try:
        # Extract data based on selected method
        if extraction_method == "auto":
            data, used_method = extract_with_ensemble(str(file_path))
            result["extraction_method"] = used_method
        elif extraction_method == "tesseract":
            data = extract_with_tesseract(str(file_path))
        elif extraction_method == "easyocr":
            data = extract_with_easyocr(str(file_path))
        elif extraction_method == "paddleocr":
            data = extract_with_paddleocr(str(file_path))
        elif extraction_method == "pdf_table":
            data = extract_from_pdf_tables(str(file_path))
        elif extraction_method == "cv_table":
            data = extract_table_with_cv(str(file_path))
        else:
            # Default to ensemble
            data, used_method = extract_with_ensemble(str(file_path))
            result["extraction_method"] = used_method
        
        # Clean and normalize data
        cleaned_data = []
        for row in data:
            # Remove empty strings and normalize whitespace
            cleaned_row = [str(cell).strip() for cell in row if str(cell).strip()]
            if cleaned_row:  # Only add non-empty rows
                cleaned_data.append(cleaned_row)
        
        # Generate preview (first few rows for UI)
        if cleaned_data:
            headers = cleaned_data[0] if len(cleaned_data) > 0 else []
            preview_rows = cleaned_data[1:6] if len(cleaned_data) > 1 else []  # Up to 5 rows for preview
            
            # Create preview as list of dicts
            preview_data = []
            for row in preview_rows:
                row_dict = {}
                for i, cell in enumerate(row):
                    header_key = headers[i] if i < len(headers) else f"Column {i+1}"
                    row_dict[header_key] = cell
                preview_data.append(row_dict)
                
            result["preview_data"] = preview_data
        
        # Save to CSV
        csv_filename = f"{task_id}.csv"
        csv_path = user_results / csv_filename
        save_to_csv(cleaned_data, str(csv_path))
        result["csv_path"] = str(csv_path)
        
        # Update status
        result["status"] = "completed"
        
    except Exception as e:
        logger.error(f"Error processing image: {str(e)}")
        result["status"] = "failed"
        result["error"] = str(e)
    
    # Calculate processing time
    result["processing_time"] = time.time() - start_time
    
    # Save result metadata
    with open(user_results / f"{task_id}.json", 'w') as f:
        json.dump(result, f)
    
    return result

# API Routes
@router.post("/{user_id}/extract", response_model=ExtractionResponse)
async def extract_from_image(
    background_tasks: BackgroundTasks,
    user_id: str = Path(..., description="User ID for storing results"),
    file: UploadFile = File(..., description="Image file to process"),
    extraction_method: str = Form("auto", description="Extraction method: auto, tesseract, easyocr, paddleocr, pdf_table, cv_table")
):
    """
    Extract data from uploaded image and convert to Excel/CSV
    
    This endpoint accepts various image types including:
    - Table images
    - Excel screenshots
    - Bill images
    - Record book images
    - Low resolution images
    
    The data will be extracted and converted to both Excel and CSV formats.
    """
    # Validate extraction method
    valid_methods = ["auto", "tesseract", "easyocr", "paddleocr", "pdf_table", "cv_table"]
    if extraction_method not in valid_methods:
        raise HTTPException(status_code=400, detail=f"Invalid extraction method. Choose from: {', '.join(valid_methods)}")
    
    # Create user directory if not exists
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    # Generate task ID
    task_id = str(uuid.uuid4())
    
    # Save uploaded file
    file_extension = os.path.splitext(file.filename)[1].lower()
    valid_extensions = ['.jpg', '.jpeg', '.png', '.tiff', '.bmp', '.pdf']
    
    if file_extension not in valid_extensions:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid file format. Supported formats: {', '.join(valid_extensions)}"
        )
    
    file_path = user_uploads / f"{task_id}{file_extension}"
    
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
    
    # Start background processing
    background_tasks.add_task(
        process_image, 
        user_id, 
        file_path,
        file.filename,
        task_id,
        extraction_method
    )
    
    return ExtractionResponse(
        task_id=task_id,
        message="Image processing started",
        status="processing"
    )

@router.get("/{user_id}/tasks/{task_id}", response_model=ExtractionResult)
async def get_task_status(user_id: str, task_id: str):
    """Get the status and results of a specific extraction task"""
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    result_file = user_results / f"{task_id}.json"
    
    if not result_file.exists():
        # Check if task exists but is still processing
        potential_uploads = list(user_uploads.glob(f"{task_id}.*"))
        if potential_uploads:
            return ExtractionResult(
                task_id=task_id,
                user_id=user_id,
                original_filename="",
                extraction_method="",
                preview_data=[],
                status="processing",
                processing_time=0
            )
        else:
            raise HTTPException(status_code=404, detail="Task not found")
    
    # Load result data
    with open(result_file, 'r') as f:
        result = json.load(f)
    
    return ExtractionResult(**result)

@router.get("/{user_id}/tasks/{task_id}/download")
async def download_result(
    user_id: str,
    task_id: str,
    format: str = Query("csv", description="Format to download: excel or csv")
):
    """Return the extracted data structure (columns and sample rows)"""
    user_uploads, user_results, user_temp = get_user_dir(user_id)

    if format.lower() == "excel":
        file_path = user_results / f"{task_id}.xlsx"
    elif format.lower() == "csv":
        file_path = user_results / f"{task_id}.csv"
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'csv'")

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Result file not found")

    # Read the file into a DataFrame
    try:
        if format.lower() == "excel":
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path, on_bad_lines="skip")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading file: {e}")
    
    sample_rows = json.loads(df.to_json(orient="records"))

    # Return structure and content
    return {
        "columns": list(df.columns),
        "sample_rows": sample_rows,
        "instructions": "This is your extracted data preview. You can download or use it for further processing."
    }

@router.get("/{user_id}/tasks")
async def list_user_tasks(user_id: str):
    """List all extraction tasks for a specific user"""
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    # Find all result JSON files
    result_files = list(user_results.glob("*.json"))
    tasks = []
    
    for result_file in result_files:
        try:
            with open(result_file, 'r') as f:
                result = json.load(f)
                tasks.append({
                    "task_id": result.get("task_id"),
                    "original_filename": result.get("original_filename"),
                    "status": result.get("status"),
                    "extraction_method": result.get("extraction_method"),
                    "processing_time": result.get("processing_time")
                })
        except Exception as e:
            logger.error(f"Error reading task result file {result_file}: {str(e)}")
    
    # Sort by most recent first (assuming task_id contains timestamp info)
    tasks.sort(key=lambda x: x["task_id"] or "", reverse=True)
    
    return {"user_id": user_id, "tasks": tasks}

@router.get("/languages")
async def get_supported_languages():
    """Get list of supported OCR languages."""
    # This is a simplified list, actual supported languages depend on Tesseract installation
    languages = {
         "afr": "Afrikaans",
        "amh": "Amharic",
        "ara": "Arabic",
        "asm": "Assamese",
        "aze": "Azerbaijani",
        "aze_cyrl": "Azerbaijani - Cyrillic",
        "bel": "Belarusian",
        "ben": "Bengali",
        "bod": "Tibetan",
        "bos": "Bosnian",
        "bul": "Bulgarian",
        "cat": "Catalan; Valencian",
        "ceb": "Cebuano",
        "ces": "Czech",
        "chi_sim": "Chinese - Simplified",
        "chi_sim_vert": "Chinese - Simplified (Vertical)",
        "chi_tra": "Chinese - Traditional",
        "chi_tra_vert": "Chinese - Traditional (Vertical)",
        "chr": "Cherokee",
        "cym": "Welsh",
        "dan": "Danish",
        "deu": "German",
        "dzo": "Dzongkha",
        "ell": "Greek, Modern (1453-)",
        "eng": "English",
        "enm": "English, Middle (1100-1500)",
        "epo": "Esperanto",
        "est": "Estonian",
        "eus": "Basque",
        "fas": "Persian",
        "fin": "Finnish",
        "fra": "French",
        "frk": "German Fraktur",
        "frm": "French, Middle (ca. 1400-1600)",
        "gle": "Irish",
        "glg": "Galician",
        "grc": "Greek, Ancient (-1453)",
        "guj": "Gujarati",
        "hat": "Haitian; Haitian Creole",
        "heb": "Hebrew",
        "hin": "Hindi",
        "hrv": "Croatian",
        "hun": "Hungarian",
        "iku": "Inuktitut",
        "ind": "Indonesian",
        "isl": "Icelandic",
        "ita": "Italian",
        "ita_old": "Italian - Old",
        "jav": "Javanese",
        "jpn": "Japanese",
        "jpn_vert": "Japanese (Vertical)",
        "kan": "Kannada",
        "kat": "Georgian",
        "kat_old": "Georgian - Old",
        "kaz": "Kazakh",
        "khm": "Central Khmer",
        "kir": "Kirghiz; Kyrgyz",
        "kor": "Korean",
        "kor_vert": "Korean (Vertical)",
        "kur": "Kurdish (Kurmanji)", # More specific name
        "lao": "Lao",
        "lat": "Latin",
        "lav": "Latvian",
        "lit": "Lithuanian",
        "mal": "Malayalam",
        "mar": "Marathi",
        "mkd": "Macedonian",
        "mlt": "Maltese",
        "mon": "Mongolian", # Added Mongolian
        "msa": "Malay",
        "mya": "Burmese",
        "nep": "Nepali",
        "nld": "Dutch; Flemish",
        "nor": "Norwegian",
        "ori": "Oriya",
        "pan": "Panjabi; Punjabi",
        "pol": "Polish",
        "por": "Portuguese",
        "pus": "Pushto; Pashto",
        "ron": "Romanian; Moldavian; Moldovan",
        "rus": "Russian",
        "san": "Sanskrit",
        "sin": "Sinhala; Sinhalese",
        "slk": "Slovak",
        "slv": "Slovenian",
        "spa": "Spanish; Castilian",
        "spa_old": "Spanish; Castilian - Old",
        "sqi": "Albanian",
        "srp": "Serbian",
        "srp_latn": "Serbian - Latin",
        "swa": "Swahili",
        "swe": "Swedish",
        "syr": "Syriac",
        "tam": "Tamil",
        "tel": "Telugu",
        "tgk": "Tajik",
        "tgl": "Tagalog", # (Filipino is often based on Tagalog)
        "tha": "Thai",
        "tir": "Tigrinya",
        "tur": "Turkish",
        "uig": "Uighur; Uyghur",
        "ukr": "Ukrainian",
        "urd": "Urdu",
        "uzb": "Uzbek",
        "uzb_cyrl": "Uzbek - Cyrillic",
        "vie": "Vietnamese",
        "yid": "Yiddish"
    }

    return {"supported_languages": languages}

@router.get("/methods")
async def list_extraction_methods():
    """List all available extraction methods"""
    methods = [
        {
            "id": "auto",
            "name": "Automatic (Ensemble)",
            "description": "Automatically selects the best extraction method based on the image type"
        },
        {
            "id": "tesseract",
            "name": "Tesseract OCR",
            "description": "Basic OCR using Tesseract, good for clean text and simple tables"
        },
        {
            "id": "easyocr",
            "name": "EasyOCR",
            "description": "Deep learning based OCR, handles various fonts and layouts"
        },
        {
            "id": "paddleocr",
            "name": "PaddleOCR",
            "description": "High accuracy OCR powered by PaddlePaddle, works well with complex layouts"
        },
        {
            "id": "pdf_table",
            "name": "PDF Table Extraction",
            "description": "Specialized for extracting tables from PDFs or images converted to PDFs"
        },
        {
            "id": "cv_table",
            "name": "Computer Vision Table",
            "description": "Uses computer vision techniques to detect table structures"
        }
    ]
    return {"extraction_methods": methods}

@router.delete("/{user_id}/tasks/{task_id}")
async def delete_task(user_id: str, task_id: str):
    """Delete a specific extraction task and its associated files"""
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    # Find all files associated with this task
    try:
        shutil.rmtree(user_uploads)
        shutil.rmtree(user_results)
        shutil.rmtree(user_temp)
    except FileNotFoundError:
        print(f"Folder not found.")
    except PermissionError:
        print(f"Permission denied while trying to delete.")
    except Exception as e:
        print(f"Error deleting folder: {e}")
    
    return {
        "task_id": task_id,
        "user_id": user_id,
        "status": "deleted"
    }

@router.post("/{user_id}/batch-extract")
async def batch_extract(
    background_tasks: BackgroundTasks,
    user_id: str,
    files: List[UploadFile] = File(...),
    extraction_method: str = Form("auto")
):
    """Process multiple images in a batch"""
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    
    # Validate extraction method
    valid_methods = ["auto", "tesseract", "easyocr", "paddleocr", "pdf_table", "cv_table"]
    if extraction_method not in valid_methods:
        raise HTTPException(status_code=400, detail=f"Invalid extraction method. Choose from: {', '.join(valid_methods)}")
    
    # Create user directory if not exists
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    batch_id = str(uuid.uuid4())
    tasks = []
    
    # Process each file
    for file in files:
        task_id = str(uuid.uuid4())
        
        # Validate file extension
        file_extension = os.path.splitext(file.filename)[1].lower()
        valid_extensions = ['.jpg', '.jpeg', '.png', '.tiff', '.bmp', '.pdf']
        
        if file_extension not in valid_extensions:
            logger.warning(f"Skipping file {file.filename} - invalid extension {file_extension}")
            continue
        
        # Save the file
        file_path = user_uploads / f"{task_id}{file_extension}"
        try:
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
        except Exception as e:
            logger.error(f"Failed to save file {file.filename}: {str(e)}")
            continue
        
        # Add background task
        background_tasks.add_task(
            process_image,
            user_id,
            file_path,
            file.filename,
            task_id,
            extraction_method
        )
        
        tasks.append({
            "task_id": task_id,
            "original_filename": file.filename,
            "status": "processing"
        })
    
    if not tasks:
        raise HTTPException(status_code=400, detail="No valid files to process")
    
    # Save batch information
    batch_info = {
        "batch_id": batch_id,
        "user_id": user_id,
        "tasks": tasks,
        "extraction_method": extraction_method,
        "timestamp": time.time()
    }
    
    with open(user_results / f"batch_{batch_id}.json", 'w') as f:
        json.dump(batch_info, f)
    
    return {
        "batch_id": batch_id,
        "tasks": tasks,
        "message": f"Started batch processing of {len(tasks)} files"
    }

@router.get("/{user_id}/batches")
async def list_user_batches(user_id: str):
    """List all batch extraction jobs for a specific user"""
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    # Find all batch JSON files
    batch_files = list(user_results.glob("batch_*.json"))
    batches = []
    
    for batch_file in batch_files:
        try:
            with open(batch_file, 'r') as f:
                batch_info = json.load(f)
                # Check if all tasks are completed
                if "tasks" in batch_info:
                    # Update task statuses based on result files
                    for task in batch_info["tasks"]:
                        task_id = task.get("task_id")
                        if task_id:
                            result_file = user_results / f"{task_id}.json"
                            if result_file.exists():
                                with open(result_file, 'r') as tf:
                                    task_result = json.load(tf)
                                    task["status"] = task_result.get("status", "unknown")
                
                batches.append({
                    "batch_id": batch_info.get("batch_id"),
                    "task_count": len(batch_info.get("tasks", [])),
                    "extraction_method": batch_info.get("extraction_method"),
                    "timestamp": batch_info.get("timestamp")
                })
        except Exception as e:
            logger.error(f"Error reading batch file {batch_file}: {str(e)}")
    
    # Sort by most recent first
    batches.sort(key=lambda x: x.get("timestamp", 0), reverse=True)
    
    return {"user_id": user_id, "batches": batches}

@router.get("/{user_id}/batches/{batch_id}")
async def get_batch_details(user_id: str, batch_id: str):
    """Get detailed information about a specific batch job"""
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    batch_file = user_results / f"batch_{batch_id}.json"
    if not batch_file.exists():
        raise HTTPException(status_code=404, detail="Batch not found")
    
    try:
        with open(batch_file, 'r') as f:
            batch_info = json.load(f)
            
        # Update task statuses
        if "tasks" in batch_info:
            for task in batch_info["tasks"]:
                task_id = task.get("task_id")
                if task_id:
                    result_file = user_results / f"{task_id}.json"
                    if result_file.exists():
                        with open(result_file, 'r') as tf:
                            task_result = json.load(tf)
                            task["status"] = task_result.get("status", "unknown")
                            # Add preview data if available
                            if "preview_data" in task_result:
                                task["preview_data"] = task_result["preview_data"]
        
        return batch_info
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading batch details: {str(e)}")

@router.post("/merge-tasks")
async def merge_extraction_results(
    user_id: str = Form(...),
    task_ids: str = Form(...),  # Comma-separated task IDs
    merge_name: str = Form(...)
):
    """Merge multiple extraction results into a single Excel/CSV file"""
    user_uploads, user_results, user_temp = get_user_dir(user_id)
    
    # Parse task IDs
    task_id_list = [tid.strip() for tid in task_ids.split(',') if tid.strip()]
    if not task_id_list:
        raise HTTPException(status_code=400, detail="No valid task IDs provided")
    
    # Generate a new task ID for the merged result
    merged_task_id = str(uuid.uuid4())
    
    try:
        # Initialize workbook for merged data
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)  # Remove default sheet
        
        all_data = []  # For CSV
        
        # Process each task
        for i, task_id in enumerate(task_id_list):
            result_file = user_results / f"{task_id}.json"
            excel_file = user_results / f"{task_id}.xlsx"
            
            if not result_file.exists() or not excel_file.exists():
                logger.warning(f"Task {task_id} results not found, skipping")
                continue
            
            # Read task info
            with open(result_file, 'r') as f:
                task_info = json.load(f)
            
            # Create a sheet name based on original filename or task index
            sheet_name = f"Sheet{i+1}"
            if "original_filename" in task_info:
                base_name = os.path.splitext(task_info["original_filename"])[0]
                # Limit sheet name to 31 chars (Excel limit)
                sheet_name = base_name[:31]
                # Replace invalid sheet name chars
                sheet_name = sheet_name.replace('[', '').replace(']', '').replace(':', '').replace('*', '')
                sheet_name = sheet_name.replace('?', '').replace('/', '').replace('\\', '')
            
            # Read data from CSV for consistent format
            csv_file = user_results / f"{task_id}.csv"
            if csv_file.exists():
                with open(csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    data = list(reader)
                    
                    # Add to Excel workbook
                    ws = wb.create_sheet(title=sheet_name)
                    for row in data:
                        ws.append(row)
                    
                    # Add to all_data for CSV with sheet name as first column
                    if data:
                        # Add header with sheet name column
                        if len(all_data) == 0:
                            headers = ["Source"] + data[0]
                            all_data.append(headers)
                        
                        # Add data rows with sheet name
                        for row in data[1:]:  # Skip header
                            all_data.append([sheet_name] + row)
        
        if len(wb.sheetnames) == 0:
            raise HTTPException(status_code=400, detail="No valid data found in any of the tasks")
        
        # Save merged Excel
        merged_excel_path = user_results / f"{merged_task_id}.xlsx"
        wb.save(str(merged_excel_path))
        
        # Save merged CSV
        merged_csv_path = user_results / f"{merged_task_id}.csv"
        with open(merged_csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            for row in all_data:
                writer.writerow(row)
        
        # Create result metadata
        merged_result = {
            "task_id": merged_task_id,
            "user_id": user_id,
            "original_filename": merge_name,
            "extraction_method": "merged",
            "excel_path": str(merged_excel_path),
            "csv_path": str(merged_csv_path),
            "preview_data": all_data[:6] if len(all_data) > 0 else [],  # First 5 rows for preview
            "status": "completed",
            "merged_from": task_id_list,
            "processing_time": 0
        }
        
        # Save result metadata
        with open(user_results / f"{merged_task_id}.json", 'w') as f:
            json.dump(merged_result, f)
        
        return {
            "task_id": merged_task_id,
            "message": f"Successfully merged {len(wb.sheetnames)} extraction results",
            "status": "completed"
        }
        
    except Exception as e:
        logger.error(f"Error merging tasks: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error merging tasks: {str(e)}")

@router.post("/preview-extraction")
async def preview_extraction_method(
    file: UploadFile = File(...),
    method: str = Form("auto")
):
    """Preview extraction results using different methods without saving"""
    # Validate method
    valid_methods = ["auto", "tesseract", "easyocr", "paddleocr", "pdf_table", "cv_table"]
    if method not in valid_methods:
        raise HTTPException(status_code=400, detail=f"Invalid method. Choose from: {', '.join(valid_methods)}")
    
    # Create temp directory
    with tempfile.TemporaryDirectory() as temp_dir:
        file_path = os.path.join(temp_dir, file.filename)
        
        # Save the uploaded file temporarily
        try:
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
        
        try:
            # Extract data based on method
            start_time = time.time()
            
            if method == "auto":
                data, used_method = extract_with_ensemble(file_path)
                method = used_method  # Update with the actual method used
            elif method == "tesseract":
                data = extract_with_tesseract(file_path)
            elif method == "easyocr":
                data = extract_with_easyocr(file_path)
            elif method == "paddleocr":
                data = extract_with_paddleocr(file_path)
            elif method == "pdf_table":
                data = extract_from_pdf_tables(file_path)
            elif method == "cv_table":
                data = extract_table_with_cv(file_path)
            
            processing_time = time.time() - start_time
            
            # Clean and normalize data
            cleaned_data = []
            for row in data:
                # Remove empty strings and normalize whitespace
                cleaned_row = [str(cell).strip() for cell in row if str(cell).strip()]
                if cleaned_row:  # Only add non-empty rows
                    cleaned_data.append(cleaned_row)
            
            # Generate preview
            preview = cleaned_data[:10]  # First 10 rows
            
            return {
                "method": method,
                "preview": preview,
                "row_count": len(cleaned_data),
                "processing_time": processing_time
            }
            
        except Exception as e:
            logger.error(f"Error in preview extraction: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Extraction error: {str(e)}")
